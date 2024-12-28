import os
import pandas as pd
from tkinter import Tk, filedialog, messagebox
from flet import *
import win32com.client as win32
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import flet

# Importando constantes do Excel
xlRowField = 1
xlDataField = 2
xlColumnField = 3
xlPageField = 4
xlSum = -4105  # Excel constant for sum


def unificar_arquivos(
    caminho_origem, caminho_resultado, linhas, valores, colunas, filtros
):
    if not caminho_origem or not caminho_resultado:
        messagebox.showwarning(
            "Aviso", "Os campos de origem e resultado devem ser preenchidos!"
        )
        return

    # Exibir mensagem de carregamento
    messagebox.showinfo("Aguarde", "O código está em execução, por favor aguarde...")

    try:
        all_data = pd.DataFrame()
        for file_name in os.listdir(caminho_origem):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(caminho_origem, file_name)
                df = pd.read_excel(file_path)
                df["Origem"] = file_name
                all_data = pd.concat([all_data, df], ignore_index=True)

        caminho_resultado_arquivo = os.path.join(
            caminho_resultado, "Dados_Unificados.xlsx"
        )

        with pd.ExcelWriter(caminho_resultado_arquivo, engine="openpyxl") as writer:
            all_data.to_excel(writer, index=False, sheet_name="Dados")

            # Formatar cabeçalho
            workbook = writer.book
            worksheet = writer.sheets["Dados"]
            header_fill = PatternFill(
                start_color="003366", end_color="003366", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            for col_num, value in enumerate(all_data.columns):
                cell = worksheet.cell(row=1, column=col_num + 1)
                cell.fill = header_fill
                cell.font = header_font

            # Linhas zebras
            for row_num in range(2, len(all_data) + 2):
                if row_num % 2 == 0:
                    for col in range(1, len(all_data.columns) + 1):
                        worksheet.cell(row=row_num, column=col).fill = PatternFill(
                            start_color="f2f2f2", end_color="f2f2f2", fill_type="solid"
                        )

            # Desabilitar linhas de grade
            worksheet.sheet_view.showGridLines = False

            # Ajustar colunas
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = adjusted_width

        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(caminho_resultado_arquivo)

        ws_info = wb.Worksheets.Add()
        ws_info.Name = "Informação"

        ws_dados = wb.Worksheets["Dados"]
        rng_tabela = ws_dados.UsedRange

        tabela_cache = wb.PivotCaches().Create(SourceType=1, SourceData=rng_tabela)
        tabela_dinamica = tabela_cache.CreatePivotTable(
            TableDestination=ws_info.Range("A1"), TableName="TabelaDinamica"
        )

        def adicionar_campo(campo, orientacao):
            try:
                if campo.strip():
                    tabela_dinamica.PivotFields(campo.strip()).Orientation = orientacao
            except Exception as e:
                print(f"Erro ao adicionar campo '{campo}': {e}")

        for campo in linhas.split(","):
            adicionar_campo(campo, xlRowField)

        for campo in valores.split(","):
            try:
                if campo.strip():
                    pf = tabela_dinamica.PivotFields(campo.strip())
                    pf.Orientation = xlDataField
                    pf.Function = xlSum
            except Exception as e:
                print(f"Erro ao adicionar campo de valor '{campo}': {e}")

        for campo in colunas.split(","):
            adicionar_campo(campo, xlColumnField)

        for campo in filtros.split(","):
            adicionar_campo(campo, xlPageField)

        ws_info.Activate()
        excel.ActiveWindow.DisplayGridlines = False

        wb.Save()
        wb.Close()
        excel.Quit()

        messagebox.showinfo(
            "Sucesso", "Unificação e Tabela Dinâmica concluídas com sucesso!"
        )

        resposta = messagebox.askyesno(
            "Abrir Arquivo", "Processo finalizado com sucesso! Deseja abrir o arquivo?"
        )
        if resposta:
            os.startfile(caminho_resultado_arquivo)

    except Exception as e:
        messagebox.showerror("Erro", f"Erro durante a unificação: {e}")


def browse_folder(target_text_field):
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        target_text_field.value = folder_selected
        target_text_field.update()
    root.destroy()


def manage(page: Page):
    page.window_max_width = 360
    page.window_width = 360
    page.window_max_height = 640
    page.window_height = 640
    page.padding = 0
    page.title = "Unificação de arquivos Excel"  # Título da janela

    # Título da seção
    title_text = Text(
        "Unificação de Planilhas", size=18, weight="bold", color="gray"
    )  # "#67b19f")
    # Subtítulo para a seção de tabela dinâmica
    subtitle_text = Text("Dados para criar tabela Dinâmica", size=14, color="gray")

    origem_input = TextField(label="Busca pasta Origem", width=250, disabled=True)
    origem_btn = IconButton(
        icon=icons.SEARCH, on_click=lambda _: browse_folder(origem_input)
    )

    resultado_input = TextField(
        label="Salva dados Unificados", width=250, disabled=True
    )
    resultado_btn = IconButton(
        icon=icons.SEARCH, on_click=lambda _: browse_folder(resultado_input)
    )

    linhas_input = TextField(label="Linhas", width=250)
    valores_input = TextField(label="Colunas", width=250)
    colunas_input = TextField(label="Filtro", width=250)
    filtros_input = TextField(label="Valores", width=250)

    unificar_btn = ElevatedButton(
        text="Unificar Arquivos",
        on_click=lambda _: unificar_arquivos(
            origem_input.value,
            resultado_input.value,
            linhas_input.value,
            valores_input.value,
            colunas_input.value,
            filtros_input.value,
        ),
    )

    body = Container(
        Column(
            [
                title_text,  # Título da seção
                Row([origem_input, origem_btn], alignment="center"),
                Row([resultado_input, resultado_btn], alignment="center"),
                subtitle_text,  # Subtítulo da seção
                Row([linhas_input], alignment="center"),
                Row([valores_input], alignment="center"),
                Row([colunas_input], alignment="center"),
                Row([filtros_input], alignment="center"),
                Row([unificar_btn], alignment="center"),
            ],
            alignment="center",
            horizontal_alignment="center",
        ),
        gradient=LinearGradient(
            begin=alignment.top_left,
            end=alignment.bottom_right,
            colors=["white", "#67b19f"],
        ),
        width=360,
        height=570,  # Ajustar altura para deixar espaço para o rodapé
        alignment=alignment.center,
    )

    # Adiciona o corpo e o rodapé na página
    page.add(body)

    # Rodapé na página
    footer = Text("EO SOLUÇÕES & CRIAÇÕES", size=12, color="gray")  # Cor do rodapé
    page.add(Row([footer], alignment="center"))  # Alinhado à esquerda


flet.app(manage)
